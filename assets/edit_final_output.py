import openpyxl
import os
from dotenv import load_dotenv
load_dotenv()


def editFinalOutput(final_checked_prices):
    xlsx_file_path = os.getenv('EDITFILE_PATH')
    workbook = openpyxl.load_workbook(xlsx_file_path)
    sheet = workbook.active
    for price in final_checked_prices:
        key = price[0]
        value = price[1]
        if len(price) == 3 :
            price = price[2]
        else :
            price = ""
        for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1, values_only=False), start=1):
            cell1 = sheet.cell(row=row_index, column=1)
            cell2 = sheet.cell(row=row_index, column=2)
            cell3 = sheet.cell(row=row_index, column=3)
            updateFile(key, value, price, cell1, cell2, cell3)
            break
    workbook.save(xlsx_file_path)


def updateFile(key, value, price, cell1, cell2, cell3) :
    cell1.value = key
    cell2.value = value
    cell3.value = price
    return